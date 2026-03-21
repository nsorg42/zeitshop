from __future__ import annotations

import csv
from datetime import date, datetime
from decimal import Decimal
from io import StringIO
from pathlib import Path
import re
from typing import Mapping, Sequence
import xml.etree.ElementTree as ET
import zipfile

from ..core.models import DiamondRecord
from ..core.normalize import normalize_text
from .detect import detect_encoding, sniff_dialect
from .xlsx_images import extract_xlsx_row_images

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:  # pragma: no cover - dependency is declared in pyproject
    load_workbook = None

_XLSX_NS_MAIN = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_XLSX_NS_REL = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
_CELL_REF_RE = re.compile(r"([A-Za-z]+)")


CANONICAL_COLUMNS = [
    "Bild",
    "Filiale",
    "Kategorie",
    "Warengruppe",
    "Marke",
    "Produktlinie",
    "Artikel Nr",
    "Kurzbeschreibung",
    "Referenz",
    "Menge",
    "Einstand",
    "Verkauf",
]


ALIASES: dict[str, str] = {
    "bild": "Bild",
    "filiale": "Filiale",
    "kategorie": "Kategorie",
    "warengruppe": "Warengruppe",
    "marke": "Marke",
    "produktlinie": "Produktlinie",
    "artikel nr": "Artikel Nr",
    "artikelnr": "Artikel Nr",
    "kurzbeschreibung": "Kurzbeschreibung",
    "referenz": "Referenz",
    "menge": "Menge",
    "einstand": "Einstand",
    "verkauf": "Verkauf",
}


def _canonical_header(raw_header: str) -> str:
    """Normalize source header and map aliases to canonical names."""
    cleaned = normalize_text(raw_header).lstrip("\ufeff")
    key = cleaned.casefold()
    return ALIASES.get(key, cleaned)


def _cell_to_text(value: object) -> str:
    """Convert CSV/XLSX cell values to normalized text."""
    if value is None:
        return ""
    if isinstance(value, str):
        return normalize_text(value)
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return normalize_text(str(value))


def _header_indexes(raw_header: Sequence[object]) -> tuple[list[int], list[str]]:
    cleaned_header = [_canonical_header(_cell_to_text(value)) for value in raw_header]
    keep_indexes = [index for index, header in enumerate(cleaned_header) if header]
    final_header = [cleaned_header[index] for index in keep_indexes]
    return keep_indexes, final_header


def _canonicalize_row(row: Mapping[str, object]) -> dict[str, str]:
    """Project arbitrary row dict into the fixed canonical DIAMOND schema."""
    canonical = {column: "" for column in CANONICAL_COLUMNS}
    for key, value in row.items():
        canonical_key = _canonical_header(key)
        if canonical_key in canonical:
            canonical[canonical_key] = _cell_to_text(value)
    return canonical


def _merge_bild_values(existing: str, image_paths: Sequence[Path]) -> str:
    values: list[str] = []
    seen: set[str] = set()

    if normalize_text(existing):
        for part in normalize_text(existing).replace("|", ";").split(";"):
            item = normalize_text(part)
            if not item:
                continue
            key = item.casefold()
            if key in seen:
                continue
            seen.add(key)
            values.append(item)

    for image_path in image_paths:
        item = str(image_path.resolve())
        key = item.casefold()
        if key in seen:
            continue
        seen.add(key)
        values.append(item)

    return "; ".join(values)


def _has_product_identity(canonical_row: Mapping[str, str]) -> bool:
    return bool(
        canonical_row.get("Artikel Nr")
        or canonical_row.get("Referenz")
        or canonical_row.get("Kurzbeschreibung")
    )


def _tokenize_header_candidate(value: str) -> str:
    """Normalize a text token for robust header-like row comparisons."""
    return re.sub(r"[^0-9a-z]+", "", normalize_text(value).casefold())


_HEADER_TOKENS = {
    column: _tokenize_header_candidate(column) for column in CANONICAL_COLUMNS
}
_IDENTITY_COLUMNS = ("Artikel Nr", "Referenz", "Kurzbeschreibung")


def _is_header_like_row(canonical_row: Mapping[str, str]) -> bool:
    """Detect repeated in-body header rows from paginated DIAMOND exports."""
    matches = 0
    identity_matches = 0

    for column, expected_token in _HEADER_TOKENS.items():
        value_token = _tokenize_header_candidate(canonical_row.get(column, ""))
        if not value_token or value_token != expected_token:
            continue
        matches += 1
        if column in _IDENTITY_COLUMNS:
            identity_matches += 1

    return identity_matches >= 1 and matches >= 3


def _parse_column_index(cell_ref: str) -> int:
    """Convert Excel cell reference (e.g. 'C10') to zero-based column index."""
    match = _CELL_REF_RE.match(cell_ref)
    if match is None:
        return -1

    letters = match.group(1).upper()
    index = 0
    for character in letters:
        index = index * 26 + (ord(character) - ord("A") + 1)
    return index - 1


def _parse_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    """Load shared string table used by XLSX `t=s` cells."""
    try:
        payload = archive.read("xl/sharedStrings.xml")
    except KeyError:
        return []

    root = ET.fromstring(payload)
    values: list[str] = []
    for item in root.findall("a:si", _XLSX_NS_MAIN):
        fragments = [node.text or "" for node in item.findall(".//a:t", _XLSX_NS_MAIN)]
        values.append("".join(fragments))
    return values


def _load_first_sheet_path(archive: zipfile.ZipFile) -> str:
    """Resolve the workbook's first sheet XML path inside the XLSX archive."""
    workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
    first_sheet = workbook_xml.find("a:sheets/a:sheet", _XLSX_NS_MAIN)
    if first_sheet is None:
        raise ValueError("XLSX workbook does not contain any sheets.")

    rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
    if not rel_id:
        raise ValueError("XLSX workbook sheet relationship is missing.")

    rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for relationship in rels_xml.findall("r:Relationship", _XLSX_NS_REL):
        if relationship.attrib.get("Id") != rel_id:
            continue
        target = relationship.attrib.get("Target", "")
        if not target:
            break
        if target.startswith("/"):
            return target.lstrip("/")
        return f"xl/{target.lstrip('/')}"

    raise ValueError("Could not resolve first worksheet path from XLSX workbook.")


def _xlsx_cell_value(cell: ET.Element, shared_strings: Sequence[str]) -> object:
    """Convert XLSX XML cell element into a Python scalar."""
    cell_type = cell.attrib.get("t")
    value_node = cell.find("a:v", _XLSX_NS_MAIN)

    if cell_type == "inlineStr":
        inline = cell.find("a:is", _XLSX_NS_MAIN)
        if inline is None:
            return ""
        fragments = [node.text or "" for node in inline.findall(".//a:t", _XLSX_NS_MAIN)]
        return "".join(fragments)

    if value_node is None:
        return ""

    raw = value_node.text or ""
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return raw
    if cell_type == "b":
        return raw == "1"
    if cell_type in {"str", "e"}:
        return raw

    # Numeric cells are plain text in XLSX XML.
    try:
        number = float(raw)
    except ValueError:
        return raw
    if number.is_integer():
        return int(number)
    return number


def _read_xlsx_rows_xml(path: Path) -> list[tuple[object, ...]]:
    """Read first worksheet rows from XLSX using stdlib XML parsing only."""
    with zipfile.ZipFile(path) as archive:
        shared_strings = _parse_shared_strings(archive)
        sheet_path = _load_first_sheet_path(archive)
        sheet_xml = ET.fromstring(archive.read(sheet_path))

    rows: list[tuple[object, ...]] = []
    for row_node in sheet_xml.findall(".//a:sheetData/a:row", _XLSX_NS_MAIN):
        try:
            row_number = int(row_node.attrib.get("r", "0"))
        except ValueError:
            row_number = 0
        if row_number <= 0:
            row_number = len(rows) + 1

        while len(rows) < row_number - 1:
            rows.append(tuple())

        values_by_index: dict[int, object] = {}
        max_index = -1
        for cell in row_node.findall("a:c", _XLSX_NS_MAIN):
            cell_ref = cell.attrib.get("r", "")
            col_index = _parse_column_index(cell_ref)
            if col_index < 0:
                continue
            values_by_index[col_index] = _xlsx_cell_value(cell, shared_strings)
            if col_index > max_index:
                max_index = col_index

        if max_index < 0:
            rows.append(tuple())
            continue

        row_values = [values_by_index.get(index, "") for index in range(max_index + 1)]
        rows.append(tuple(row_values))

    return rows


def read_diamond_csv(path: str | Path) -> list[DiamondRecord]:
    """Read a DIAMOND CSV file into canonical `DiamondRecord` objects."""

    file_path = Path(path)
    raw_bytes = file_path.read_bytes()
    encoding = detect_encoding(raw_bytes)
    text = raw_bytes.decode(encoding, errors="replace")

    sample = text[:4096]
    dialect = sniff_dialect(sample)

    stream = StringIO(text)
    reader = csv.reader(stream, dialect=dialect)

    try:
        raw_header = next(reader)
    except StopIteration:
        return []

    keep_indexes, final_header = _header_indexes(raw_header)

    records: list[DiamondRecord] = []
    for source_row, raw_row in enumerate(reader, start=2):
        if not any(_cell_to_text(cell) for cell in raw_row):
            continue

        row_values = [raw_row[index] if index < len(raw_row) else "" for index in keep_indexes]
        row_map = dict(zip(final_header, row_values, strict=False))
        canonical = _canonicalize_row(row_map)

        # Ignore footer rows that carry totals but no product identity.
        if not _has_product_identity(canonical):
            continue
        if _is_header_like_row(canonical):
            continue

        records.append(DiamondRecord(source_row=source_row, data=canonical))

    return records


def _find_header_row(rows: Sequence[Sequence[object]]) -> tuple[int, list[int], list[str]]:
    """Find the row most likely to be the DIAMOND header in XLSX exports."""
    best_row = 0
    best_score = 0
    best_keep_indexes: list[int] = []
    best_final_header: list[str] = []

    for row_index, raw_row in enumerate(rows, start=1):
        keep_indexes, final_header = _header_indexes(raw_row)
        if not final_header:
            continue

        score = sum(1 for value in final_header if value in CANONICAL_COLUMNS)
        has_identity = any(
            value in {"Artikel Nr", "Referenz", "Kurzbeschreibung"}
            for value in final_header
        )
        if not has_identity:
            continue

        if score > best_score:
            best_row = row_index
            best_score = score
            best_keep_indexes = keep_indexes
            best_final_header = final_header

    if best_score == 0:
        raise ValueError(
            "Could not locate a DIAMOND header row in XLSX file. "
            "Expected columns like 'Artikel Nr', 'Referenz', or 'Kurzbeschreibung'."
        )

    return best_row, best_keep_indexes, best_final_header


def read_diamond_xlsx(
    path: str | Path,
    *,
    extract_embedded_images: bool = False,
    image_export_dir: str | Path | None = None,
) -> list[DiamondRecord]:
    """Read a DIAMOND XLSX file into canonical `DiamondRecord` objects."""
    file_path = Path(path)
    if load_workbook is None:
        rows = _read_xlsx_rows_xml(file_path)
    else:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

    if not rows:
        return []

    header_row, keep_indexes, final_header = _find_header_row(rows)

    records: list[DiamondRecord] = []
    data_rows = rows[header_row:]
    for source_row, raw_row in enumerate(data_rows, start=header_row + 1):
        if not any(_cell_to_text(cell) for cell in raw_row):
            continue

        row_values = [raw_row[index] if index < len(raw_row) else "" for index in keep_indexes]
        row_map = dict(zip(final_header, row_values, strict=False))
        canonical = _canonicalize_row(row_map)

        if not _has_product_identity(canonical):
            continue
        if _is_header_like_row(canonical):
            continue

        records.append(DiamondRecord(source_row=source_row, data=canonical))

    if not extract_embedded_images or not records:
        return records

    with zipfile.ZipFile(file_path) as archive:
        sheet_path = _load_first_sheet_path(archive)

    row_images = extract_xlsx_row_images(
        file_path,
        sheet_path=sheet_path,
        header_row=header_row,
        keep_indexes=keep_indexes,
        final_header=final_header,
        source_rows=[record.source_row for record in records],
        output_dir=image_export_dir,
    )
    if not row_images:
        return records

    merged_records: list[DiamondRecord] = []
    for record in records:
        image_paths = row_images.get(record.source_row, [])
        if not image_paths:
            merged_records.append(record)
            continue

        merged_data = dict(record.data)
        merged_data["Bild"] = _merge_bild_values(merged_data.get("Bild", ""), image_paths)
        merged_records.append(DiamondRecord(source_row=record.source_row, data=merged_data))

    return merged_records


def read_diamond_file(
    path: str | Path,
    *,
    extract_embedded_images: bool = False,
    image_export_dir: str | Path | None = None,
) -> list[DiamondRecord]:
    """Read DIAMOND exports from CSV or XLSX."""
    file_path = Path(path)
    suffix = file_path.suffix.casefold()
    if suffix == ".csv":
        return read_diamond_csv(file_path)
    if suffix == ".xlsx":
        return read_diamond_xlsx(
            file_path,
            extract_embedded_images=extract_embedded_images,
            image_export_dir=image_export_dir,
        )

    raise ValueError(
        f"Unsupported DIAMOND format '{file_path.suffix}'. "
        "Please provide a .csv or .xlsx file."
    )
