from pathlib import Path

import pytest
from openpyxl import Workbook

from tests._xlsx_factory import build_sample_diamond_xlsx
from zeitshop_converter.io import diamond_reader
from zeitshop_converter.io.diamond_reader import read_diamond_csv
from zeitshop_converter.io.diamond_reader import read_diamond_file
from zeitshop_converter.io.diamond_reader import read_diamond_xlsx


def test_reader_keeps_bild_and_drops_empty_columns(tmp_path: Path) -> None:
    file_path = tmp_path / "diamond.csv"
    file_path.write_text(
        "Bild;Filiale;Kategorie;Warengruppe;Marke;;Produktlinie;Artikel Nr;Kurzbeschreibung;;Referenz;Menge;Einstand;;Verkauf\n"
        "images/sku-1001.jpg;Store-West;Category-A;Segment-B;Brand-X;;Linea;SKU-1001;Linea;;REF-1001;1;1'775.00;;3'550.00\n",
        encoding="cp1252",
    )

    records = read_diamond_csv(file_path)

    assert len(records) == 1
    row = records[0]
    assert row.source_row == 2
    assert row.data["Bild"] == "images/sku-1001.jpg"
    assert row.data["Filiale"] == "Store-West"
    assert row.data["Artikel Nr"] == "SKU-1001"
    assert row.data["Verkauf"] == "3'550.00"


def test_reader_parses_xlsx_with_preamble_and_header_row(tmp_path: Path) -> None:
    file_path = tmp_path / "diamond.xlsx"

    workbook = Workbook()
    worksheet = workbook.active

    worksheet["A2"] = "Lager"
    worksheet["A7"] = "Kriterien:"
    worksheet["A9"] = "Bild"
    worksheet["E9"] = "Filiale"
    worksheet["F9"] = "Artikel Nr"
    worksheet["G9"] = "Kategorie"
    worksheet["H9"] = "Warengruppe"
    worksheet["J9"] = "Marke"
    worksheet["K9"] = "Produktlinie"
    worksheet["L9"] = "Referenz"
    worksheet["N9"] = "Kurzbeschreibung"
    worksheet["O9"] = "Menge"
    worksheet["P9"] = "Einstand"
    worksheet["Q9"] = "Verkauf"

    worksheet["E11"] = "Store-West"
    worksheet["F11"] = "SKU-1001"
    worksheet["G11"] = "Category-A"
    worksheet["H11"] = "Segment-B"
    worksheet["J11"] = "Brand-X"
    worksheet["K11"] = "Linea"
    worksheet["L11"] = "REF-1001"
    worksheet["N11"] = "Linea"
    worksheet["O11"] = 1
    worksheet["P11"] = "1'775.00"
    worksheet["Q11"] = "3'550.00"

    workbook.save(file_path)

    records = read_diamond_xlsx(file_path)

    assert len(records) == 1
    row = records[0]
    assert row.source_row == 11
    assert row.data["Filiale"] == "Store-West"
    assert row.data["Artikel Nr"] == "SKU-1001"
    assert row.data["Referenz"] == "REF-1001"
    assert row.data["Menge"] == "1"
    assert row.data["Verkauf"] == "3'550.00"


def test_reader_dispatches_by_extension(tmp_path: Path) -> None:
    csv_path = tmp_path / "diamond.CSV"
    csv_path.write_text(
        "Filiale;Artikel Nr;Kurzbeschreibung;Menge;Einstand;Verkauf\n"
        "Store-West;10;Sample;1;5;9\n",
        encoding="utf-8",
    )

    csv_records = read_diamond_file(csv_path)
    assert len(csv_records) == 1
    assert csv_records[0].data["Artikel Nr"] == "10"

    xlsx_path = tmp_path / "diamond.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Artikel Nr"
    worksheet["B1"] = "Kurzbeschreibung"
    worksheet["C1"] = "Menge"
    worksheet["D1"] = "Einstand"
    worksheet["E1"] = "Verkauf"
    worksheet["A2"] = "11"
    worksheet["B2"] = "Sample X"
    worksheet["C2"] = "2"
    worksheet["D2"] = "7"
    worksheet["E2"] = "13"
    workbook.save(xlsx_path)

    xlsx_records = read_diamond_file(xlsx_path)
    assert len(xlsx_records) == 1
    assert xlsx_records[0].data["Artikel Nr"] == "11"


def test_reader_csv_empty_file_returns_no_records(tmp_path: Path) -> None:
    file_path = tmp_path / "empty.csv"
    file_path.write_text("", encoding="utf-8")

    assert read_diamond_csv(file_path) == []


def test_reader_skips_repeated_header_rows_in_csv(tmp_path: Path) -> None:
    file_path = tmp_path / "diamond.csv"
    file_path.write_text(
        "Filiale;Artikel Nr;Kurzbeschreibung;Referenz;Menge;Einstand;Verkauf\n"
        "Filiale;Artikel Nr;Kurzbeschreibung;Referenz;Menge;Einstand;Verkauf\n"
        "Store-West;10;Sample;REF-10;1;5;9\n",
        encoding="utf-8",
    )

    records = read_diamond_csv(file_path)

    assert len(records) == 1
    assert records[0].source_row == 3
    assert records[0].data["Artikel Nr"] == "10"


def test_reader_skips_repeated_header_rows_in_xlsx(tmp_path: Path) -> None:
    file_path = tmp_path / "diamond.xlsx"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Artikel Nr", "Kurzbeschreibung", "Referenz", "Menge", "Einstand", "Verkauf"])
    worksheet.append(["11", "Sample A", "REF-11", 1, "5", "9"])
    worksheet.append(["Artikel Nr", "Kurzbeschreibung", "Referenz", "Menge", "Einstand", "Verkauf"])
    worksheet.append(["12", "Sample B", "REF-12", 2, "6", "10"])
    workbook.save(file_path)

    records = read_diamond_xlsx(file_path)

    assert len(records) == 2
    assert records[0].source_row == 2
    assert records[0].data["Artikel Nr"] == "11"
    assert records[1].source_row == 4
    assert records[1].data["Artikel Nr"] == "12"


def test_reader_rejects_unsupported_file_extension(tmp_path: Path) -> None:
    file_path = tmp_path / "diamond.txt"
    file_path.write_text("test", encoding="utf-8")

    with pytest.raises(ValueError, match="Unsupported DIAMOND format"):
        read_diamond_file(file_path)


def test_reader_formats_scalar_cell_types() -> None:
    assert diamond_reader._cell_to_text(True) == "TRUE"
    assert diamond_reader._cell_to_text(2.5) == "2.5"
    assert diamond_reader._cell_to_text(3) == "3"
    assert diamond_reader._cell_to_text("  a  b  ") == "a b"


def test_reader_rejects_xlsx_without_identifiable_header() -> None:
    with pytest.raises(ValueError, match="Could not locate a DIAMOND header row"):
        diamond_reader._find_header_row([["Foo", "Bar"], ["1", "2"]])


def test_reader_xlsx_fallback_works_without_openpyxl(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(diamond_reader, "load_workbook", None)
    workbook_path = build_sample_diamond_xlsx(tmp_path / "sample_reader.xlsx")

    records = diamond_reader.read_diamond_xlsx(workbook_path)

    assert records
    assert records[0].source_row == 11
    assert records[0].data["Artikel Nr"] == "SKU-1201"
